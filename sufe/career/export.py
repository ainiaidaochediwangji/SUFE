from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from sufe.storage import ensure_directory, write_csv, write_text

SUMMARY_FIELDS = [
    "zpxxid",
    "fbrq",
    "zpzt",
    "dwmc",
    "hyyjmc",
    "xzyjmc",
    "rsgmmc",
    "szssmc",
    "szsmc",
    "szxmc",
    "xxdz",
    "zpjzrq",
    "jltdyx",
    "zpxxwz",
    "dwwz",
    "xqrs",
    "zws",
    "attachment_count",
    "has_positions",
    "detail_url",
]

SUMMARY_HEADERS = [
    "招聘信息ID",
    "发布日期",
    "招聘标题",
    "单位名称",
    "行业",
    "性质",
    "规模",
    "省份",
    "城市",
    "区县",
    "详细地址",
    "截止日期",
    "简历投递邮箱",
    "招聘信息网址",
    "单位网址",
    "需求人数",
    "职位数",
    "附件数",
    "是否有岗位",
    "详情链接",
]

POSITION_FIELDS = [
    "zpxxid",
    "fbrq",
    "dwmc",
    "zpzt",
    "gwmc",
    "xqrs",
    "gzdd",
    "xlyq",
    "zwyx",
    "zyyq",
    "yxq",
    "zwms",
    "detail_url",
]

POSITION_HEADERS = [
    "招聘信息ID",
    "发布日期",
    "单位名称",
    "招聘标题",
    "岗位名称",
    "需求人数",
    "工作地点",
    "学历要求",
    "职位月薪",
    "专业要求",
    "有效期",
    "岗位描述",
    "详情链接",
]

ATTACHMENT_FIELDS = ["zpxxid", "zpzt", "dwmc", "file_name", "attach_url"]
ATTACHMENT_HEADERS = ["招聘信息ID", "招聘标题", "单位名称", "文件名", "附件链接"]


def sort_key(value: dict[str, Any]) -> tuple[str, str]:
    return str(value.get("fbrq") or ""), str(value.get("zpxxid") or "")


def parse_attachments(value: Any) -> list[dict[str, str]]:
    if not value:
        return []
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    if not isinstance(value, str):
        return []
    try:
        data = json.loads(value)
    except json.JSONDecodeError:
        return []
    return [item for item in data if isinstance(item, dict)]


def normalize_record(item: dict[str, Any]) -> dict[str, Any]:
    record = dict(item)
    record["attachments"] = parse_attachments(record.get("zpxxAttach"))
    record["detail_url"] = f"https://career.sufe.edu.cn/career/zpxx/view/sxzpxx/{record['zpxxid']}"
    return record


def summary_row(record: dict[str, Any]) -> dict[str, Any]:
    attachments = record.get("attachments") or []
    return {
        "zpxxid": record.get("zpxxid"),
        "fbrq": record.get("fbrq"),
        "zpzt": record.get("zpzt"),
        "dwmc": record.get("dwmc"),
        "hyyjmc": record.get("hyyjmc"),
        "xzyjmc": record.get("xzyjmc"),
        "rsgmmc": record.get("rsgmmc"),
        "szssmc": record.get("szssmc"),
        "szsmc": record.get("szsmc"),
        "szxmc": record.get("szxmc"),
        "xxdz": record.get("xxdz"),
        "zpjzrq": record.get("zpjzrq"),
        "jltdyx": record.get("jltdyx"),
        "zpxxwz": record.get("zpxxwz"),
        "dwwz": record.get("dwwz"),
        "xqrs": record.get("xqrs"),
        "zws": record.get("zws"),
        "attachment_count": len(attachments),
        "has_positions": int(bool(record.get("zwxxList"))),
        "detail_url": record.get("detail_url"),
    }


def attachment_rows(record: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in record.get("attachments") or []:
        attach_url = str(item.get("attachUrl") or "")
        rows.append(
            {
                "zpxxid": record.get("zpxxid"),
                "zpzt": record.get("zpzt"),
                "dwmc": record.get("dwmc"),
                "file_name": item.get("fileName"),
                "attach_url": f"https://career.sufe.edu.cn/career{attach_url}" if attach_url.startswith("/") else attach_url,
            }
        )
    return rows


def position_rows(record: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in record.get("zwxxList") or []:
        if not isinstance(item, dict):
            continue
        rows.append(
            {
                "zpxxid": record.get("zpxxid"),
                "fbrq": record.get("fbrq"),
                "dwmc": record.get("dwmc"),
                "zpzt": record.get("zpzt"),
                "gwmc": item.get("gwmc"),
                "xqrs": item.get("xqrs") or item.get("xqrsmc"),
                "gzdd": item.get("gzdd"),
                "xlyq": item.get("xlyq") or item.get("xlyqmc"),
                "zwyx": item.get("zwyx"),
                "zyyq": item.get("zyyq"),
                "yxq": item.get("yxq"),
                "zwms": item.get("zwms"),
                "detail_url": record.get("detail_url"),
            }
        )
    return rows


def _write_excel_sheet(worksheet: Any, rows: list[dict[str, Any]], fields: list[str], headers: list[str]) -> None:
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        worksheet.append([row.get(field) for field in fields])
    worksheet.freeze_panes = "A2"
    for index, field in enumerate(fields, start=1):
        values = [field]
        for row in rows:
            value = row.get(field)
            values.append("" if value is None else str(value))
        width = min(max(len(value) for value in values) + 2, 60)
        worksheet.column_dimensions[get_column_letter(index)].width = width


def save_tables(
    output_dir: Path,
    summary_data: list[dict[str, Any]],
    position_data: list[dict[str, Any]],
    attachment_data: list[dict[str, Any]],
) -> None:
    write_csv(output_dir / "records.csv", summary_data, SUMMARY_FIELDS, SUMMARY_HEADERS)
    write_csv(output_dir / "positions.csv", position_data, POSITION_FIELDS, POSITION_HEADERS)
    write_csv(output_dir / "attachments.csv", attachment_data, ATTACHMENT_FIELDS, ATTACHMENT_HEADERS)

    ensure_directory(output_dir)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "招聘信息"
    _write_excel_sheet(summary_sheet, summary_data, SUMMARY_FIELDS, SUMMARY_HEADERS)

    position_sheet = workbook.create_sheet("岗位信息")
    _write_excel_sheet(position_sheet, position_data, POSITION_FIELDS, POSITION_HEADERS)

    attachment_sheet = workbook.create_sheet("附件信息")
    _write_excel_sheet(attachment_sheet, attachment_data, ATTACHMENT_FIELDS, ATTACHMENT_HEADERS)

    excel_path = output_dir / "career_data.xlsx"
    temp_path = excel_path.with_suffix(".tmp.xlsx")
    workbook.save(temp_path)
    workbook.close()
    os.replace(temp_path, excel_path)


def comparable(record: dict[str, Any]) -> str:
    payload = {key: value for key, value in record.items() if key != "_list_snapshot"}
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def alerts(previous: list[dict[str, Any]], current: list[dict[str, Any]]) -> dict[str, Any]:
    previous_map = {str(item["zpxxid"]): item for item in previous if item.get("zpxxid")}
    current_map = {str(item["zpxxid"]): item for item in current if item.get("zpxxid")}

    new_ids = [zpxxid for zpxxid in current_map if zpxxid not in previous_map]
    removed_ids = [zpxxid for zpxxid in previous_map if zpxxid not in current_map]
    updated_ids = [
        zpxxid
        for zpxxid in current_map
        if zpxxid in previous_map and comparable(previous_map[zpxxid]) != comparable(current_map[zpxxid])
    ]

    def rows(ids: list[str], source: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
        return [
            {
                "zpxxid": zpxxid,
                "fbrq": source[zpxxid].get("fbrq"),
                "zpzt": source[zpxxid].get("zpzt"),
                "dwmc": source[zpxxid].get("dwmc"),
                "detail_url": source[zpxxid].get("detail_url"),
            }
            for zpxxid in ids
        ]

    return {
        "new_count": len(new_ids),
        "updated_count": len(updated_ids),
        "removed_count": len(removed_ids),
        "new_records": rows(new_ids, current_map),
        "updated_records": rows(updated_ids, current_map),
        "removed_records": rows(removed_ids, previous_map),
    }


def write_notice(path: Path, change_alerts: dict[str, Any]) -> None:
    lines = [
        f"new={change_alerts['new_count']}",
        f"updated={change_alerts['updated_count']}",
        f"removed={change_alerts['removed_count']}",
    ]
    for key, title in (
        ("new_records", "new"),
        ("updated_records", "updated"),
        ("removed_records", "removed"),
    ):
        rows = change_alerts[key]
        if not rows:
            continue
        lines.append(f"[{title}]")
        for item in rows[:20]:
            lines.append(f"{item['fbrq']} | {item['dwmc']} | {item['zpzt']} | {item['zpxxid']}")
    write_text(path, "\n".join(lines) + "\n")
